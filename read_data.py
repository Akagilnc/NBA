import json
from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join


class ReadData:

    def read_file_name(self):
        files = [f for f in listdir(".\data") if isfile(join('.\data', f))]
        row_num = 1
        for filename in files:
            print('begin to make ' + str(filename))
            row_num = self.make_datas(filename, row_num)
            print('end of ' + str(filename))

        print('done')

    @classmethod
    def make_datas(cls, name, row_num):
        with open('./data/' + name) as data_file:
            data = json.load(data_file)
            data_title = data.get('name')
            data_value = data.get('value')

            wb = load_workbook('NBA.xlsx')
            sheet = wb.active
            # sheet.title = data_title
            row_num += 1

            for data in data_value:
                row_num += 1
                sheet.cell(column=1, row=row_num, value=data_title)
                sheet.cell(column=2, row=row_num, value=data['x'])
                sheet.cell(column=3, row=row_num, value=data['y'])
                if 'p' in data.keys():
                    sheet.cell(column=5, row=row_num, value=data['p'])
                if 'p' in data.keys() and data['p'] != 'f':
                    sheet.cell(column=4, row=row_num, value=data['t'])
                    sheet.cell(column=6, row=row_num, value=data['o'])
                    sheet.cell(column=7, row=row_num, value=data['d'])

            wb.save('NBA.xlsx')
            print(data_title)
        return row_num


process = ReadData()
process.read_file_name()
